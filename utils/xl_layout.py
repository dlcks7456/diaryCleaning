import openpyxl as xl
from features.setting import get_column_name
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles import Alignment

def set_xl_layout(xl_path, df, error_columns, check_columns, do_not_modify_columns):
    columns = df.columns
    panel_no = get_column_name('panel_no')
    unique_panels = df[panel_no].unique().tolist()

    for error_column in [*error_columns, *check_columns]:
        # boolean 값을 먼저 문자열로 변환하여 dtype 호환성 경고 방지
        df[error_column] = df[error_column].astype(str)
        df.loc[df[error_column] == 'False', error_column] = ''
        if error_column in check_columns :
            df.loc[df[error_column] == 'True', error_column] = '△'
        else :
            df.loc[df[error_column] == 'True', error_column] = 'X'

    df.to_excel(xl_path, sheet_name='Raw data', index=False)

    load_xl = xl.load_workbook(xl_path)
    sheet = load_xl.active

    # PANELNO 컬럼의 인덱스 찾기
    panel_col_idx = None
    if panel_no in columns:
        panel_col_idx = list(columns).index(panel_no) + 1  # 1-based index

    # 틀고정 설정 - PANELNO 컬럼과 첫번째 행 고정
    if panel_col_idx:
        # PANELNO 컬럼 다음 열과 두번째 행에서 틀고정
        freeze_cell = sheet.cell(row=2, column=panel_col_idx + 1)
        sheet.freeze_panes = freeze_cell
    else:
        # PANELNO 컬럼이 없는 경우 첫번째 행만 고정
        sheet.freeze_panes = 'A2'

    # 데이터 범위 확인
    max_row = sheet.max_row
    max_col = sheet.max_column

    table_range = f"A1:{sheet.cell(max_row, max_col).coordinate}"
    table = Table(displayName="ConvertedDataTable", ref=table_range)

    # 헤더 스타일 - #24aadf 배경색, 흰색 글씨
    header_fill = PatternFill(start_color="24aadf", end_color="24aadf", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    do_not_modify_font = Font(color="000000", bold=True, size=11)  # 검은색 글씨 폰트

    # 패널별 스타일 - 흰색과 연한 회색
    panel_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 첫번째 패널(흰색)
    panel_fill_2 = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")  # 두번째 패널(진한 회색)

    # 에러 셀 스타일 - 연한 빨간색 (핑크색)
    error_fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
    check_fill = PatternFill(start_color="ffffcc", end_color="ffffcc", fill_type="solid")
    do_not_modify_fill = PatternFill(start_color="ffd6a3", end_color="ffd6a3", fill_type="solid")  # 연한 오렌지색
    center_alignment = Alignment(horizontal='center')

    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # 에러 컬럼의 열 인덱스 찾기
    error_col_indices = []
    check_col_indices = []
    do_not_modify_col_indices = []

    for error_col in error_columns :
        if error_col in columns:
            col_idx = list(columns).index(error_col) + 1  # 1-based index
            error_col_indices.append(col_idx)

    for check_col in check_columns :
        if check_col in columns:
            col_idx = list(columns).index(check_col) + 1  # 1-based index
            check_col_indices.append(col_idx)

    for do_not_modify_col in do_not_modify_columns :
        if do_not_modify_col in columns:
            col_idx = list(columns).index(do_not_modify_col) + 1  # 1-based index
            do_not_modify_col_indices.append(col_idx)

    # 헤더 행에 스타일 적용
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        # 수정 금지 컬럼인 경우 do_not_modify_fill 적용, 아니면 header_fill 적용
        if col in do_not_modify_col_indices:
            cell.fill = do_not_modify_fill
            cell.font = do_not_modify_font  # 검은색 글씨
        else:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = thin_border

    # 패널별 색상 매핑 생성
    panel_color_map = {}
    if panel_col_idx:
        for i, panel in enumerate(unique_panels):
            panel_color_map[panel] = panel_fill_1 if i % 2 == 0 else panel_fill_2

    # 데이터 행에 스타일 적용 (패널별 구분 및 에러 셀 처리)
    for row in range(2, max_row + 1):
        # 현재 행의 패널 번호 가져오기
        current_panel = None
        if panel_col_idx:
            current_panel = sheet.cell(row=row, column=panel_col_idx).value

        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)

            # 먼저 패널별 배경색 적용
            if current_panel and current_panel in panel_color_map:
                cell.fill = panel_color_map[current_panel]
            else:
                # 기본값으로 흰색 적용
                cell.fill = panel_fill_1

            # 에러 컬럼인 경우 에러 셀 스타일로 덮어쓰기
            if col in error_col_indices or col in check_col_indices:
                cell.alignment = center_alignment
                if col in error_col_indices and cell.value == 'X':
                    cell.fill = error_fill
                elif col in check_col_indices and cell.value == '△':
                    cell.fill = check_fill

            cell.border = thin_border

    # 컬럼 너비 자동 조정
    for col in range(1, max_col + 1):
        column_letter = xl.utils.get_column_letter(col)
        max_length = 0

        # 헤더와 모든 데이터 행에서 최대 길이 찾기
        for row in range(1, max_row + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                # 한글과 영문을 구분해서 길이 계산 (한글은 2배로 계산)
                str_value = str(cell_value)
                length = sum(2 if ord(char) > 127 else 1 for char in str_value)
                max_length = max(max_length, length)

        # 최소 너비 10, 최대 너비 50으로 제한하고 여유분 추가
        adjusted_width = min(max(max_length + 2, 10), 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 기본 테이블 스타일 설정 (스트라이프 효과는 비활성화)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,  # 커스텀 스타일을 적용했으므로 비활성화
        showColumnStripes=False
    )
    table.tableStyleInfo = style

    # 시트에 테이블 추가
    sheet.add_table(table)

    # 워크북 저장
    load_xl.save(xl_path)